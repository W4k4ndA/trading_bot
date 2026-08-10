"""_summary_
"""


import asyncio
import os
import locale
import logging

from datetime import datetime, timedelta, date, time
from time import time
from colorama import init, Fore, Style, Back
from telethon import TelegramClient, events
from playwright.async_api import async_playwright, Page
from playwright import async_api
from openpyxl import load_workbook, Workbook


import variables as var


# Variable global para indicar si la función montar_operacion está en ejecución
operation_running = False

browser_page = None

telegram_client = None
broker_signal_message = ""

duracion = None
par = None
fechaHora_iniOperacion = ""
fechaHora_finOperacion = ""
operacion = ""

init_balance = None
final_balance = None

# Configurar el nivel de logging, el nombre del archivo y el formato de los mensajes
logging.basicConfig(level=logging.DEBUG, filename='logs/registro.log',
                    format='%(asctime)s - %(levelname)s\n%(message)s\n\n')

# Crear un logger
logger = logging.getLogger()

# Establecer el idioma a español de España
locale.setlocale(locale.LC_ALL, 'es_ES.utf8')


async def launch_montar_operacion_every_10_minutes():
    """Launch the 'montar_operacion' function at 10-minute intervals.

    This function runs indefinitely and triggers the 'montar_operacion'
    function every 10 minutes.

    Args:
        None

    Returns:
        None
    """
    global browser_page
    page_to_use = browser_page
    global duracion
    global par
    global direccion
    global monto
    global porcentaje

    # Loop indefinitely to schedule operations every 10 minutes
    while True:
        # Execute the operation mounting function
        await montar_operacion(page_to_use, var.broker_url_demo, 1, "USD/TRY", "PUT", 10)
        await asyncio.sleep(var.sleep_seconds_zero)
        await ejecutar_operacion(page_to_use, "put")
        # Wait for 10 minutes before the next execution
        await asyncio.sleep(900)  # 10 minutes in seconds


# This function refreshes the given context web pages until certain conditions are met
async def refresh_page(page):
    """Refresh the given pages until certain conditions are met

    Args:
        *pages: Page objects to refresh
    """
    try:
        while True:
            global operation_running
            op_running = operation_running

            # Check if the mount_operation function is not running
            if not op_running:
                error_flag = False
                error = ""

                # Reload the page and wait for it to load
                await page.reload(wait_until="load")
                await page.wait_for_load_state()
                if page.url == var.broker_url_demo or page.url == var.broker_url:
                    await page.wait_for_load_state()
                    # await asyncio.sleep(var.sleep_seconds_first)
                    await page.wait_for_selector(".section-deal__time", state="visible", timeout=10000)
                    # Check for the presence of the field to indicate the duration of the operation
                    if not await page.locator(".section-deal__time").is_visible():
                        error_flag = True
                        error = "The field to indicate the duration of the operation was not found"
                        print(error)
                    # Check for the presence of the field to indicate the amount of the operation
                    elif not await page.get_by_role("textbox").nth(1).is_visible():
                        error_flag = True
                        error = "The field to indicate the amount of the operation was not found"
                        print(error)
                    # Check for the presence of a currency pair on the page
                    elif not await page.locator("#tab-active").is_visible():
                        error_flag = True
                        error = "No currency pair was found on the page"
                        print(error)
                    # Check for the presence of the element that shows the available balance
                    elif not await page.locator(".usermenu__info-balance").is_visible():
                        error_flag = True
                        error = "The element that shows the available balance was not found"
                        print(error)

                    if error_flag:
                        await asyncio.sleep(var.sleep_seconds_second)
                        await refresh_page(page)

                print(
                    f"Página actualizada el {datetime.now().strftime(var.formato_fecha)}")
                logger.info("Página actualizada")
            # Wait x minutes before checking again
            await asyncio.sleep(60*var.refreshing_time_minutes)
    except Exception as e:
        print(
            f"Error al actualizar la pagina: {e}\nTraceback:{e.__context__}")
        logger.error("Error al actualizar la pagina: %s\nTraceback: %s",
                     e, e.__traceback__.__repr__())
        await refresh_page(page)

# Funcion para guardar los resultados de la operacion en un excel


def agregar_a_excel(fecha_operacion: date, hora_operacion: time, duracion: int, tipo_resultado: str, par_moneda: str,
                    direccion: str, balance_inicial: float, balance_final: float, profit: float, nombre_archivo: str = var.direccion_excel):
    """
    Adds data to an Excel file for a trading operation.

    Args:
    fecha_operacion (date): The date of the operation.
    hora_operacion (time): The time of the operation.
    duracion (int): The duration (in minutes) of the operation.
    tipo_resultado (str): The result type of the operation (D=Direct, G1=Gale 1, G2=Gale 2, P=Lose).
    par_moneda (str): The currency pair of the operation.
    direccion (str): The direction of the operation (PUT, CALL).
    balance_inicial (float): The initial balance before the operation.
    balance_final (float): The final balance after the operation.
    profit (float): The profit from the operation.
    nombre_archivo (str): The name of the Excel file to add the data to.

    Returns:
    None
    """

    # # Llamar a la función para agregar datos al archivo de Excel
    # try:
    #     if intentos == 0:
    #         tipo_operacion = "D"
    #     elif intentos == 1:
    #         tipo_operacion = "G1"
    #     # elif intentos == 2:
    #     #     tipo_operacion = "G2"
    #     else:
    #         tipo_operacion = "P"

    #     await asyncio.sleep(var.sleep_seconds_zero)
    #     final_balance = await page.locator(".usermenu__info-balance").inner_text()
    #     final_balance = float(final_balance[1:].replace(",", ""))

    #     agregar_a_excel(fecha_inicio, fecha_fin, tipo_operacion,
    #                     par, direc_operacion, init_balance, final_balance)
    #     print(Fore.BLUE + Style.BRIGHT +
    #         "Datos agregados al archivo de Excel exitosamente.")
    #     logger.info("Datos agregados al archivo de Excel exitosamente.")

    attemps = 0
    # print(Fore.WHITE + Style.NORMAL + "Esperando próxima señal")
    while attemps < var.max_retry_attempts:
        try:

            # Comprobar si el archivo ya existe
            if os.path.exists(nombre_archivo):
                # Cargar el archivo existente
                wb = load_workbook(nombre_archivo)
                # Seleccionar la hoja de Excel
                ws = wb.active
                # Encontrar la primera fila vacía
                fila_vacia = ws.max_row + 1
            else:
                # Si el archivo no existe, crear uno nuevo
                wb = Workbook()
                ws = wb.active
                # Escribir los encabezados en la primera fila
                ws.append(["Fecha Operación", "Hora Operación", "Duracion", "Par",
                           "Dirección", "Tipo de Resultado", "Balance Inicial", "Balance Final", "Profit"])
                # La primera fila vacía es la segunda fila
                fila_vacia = 2

            # Agregar los datos en la fila vacía
            ws.cell(row=fila_vacia, column=1, value=fecha_operacion)
            ws.cell(row=fila_vacia, column=2, value=hora_operacion)
            ws.cell(row=fila_vacia, column=3, value=duracion)
            ws.cell(row=fila_vacia, column=4, value=par_moneda)
            ws.cell(row=fila_vacia, column=5, value=direccion)
            ws.cell(row=fila_vacia, column=6, value=tipo_resultado)
            ws.cell(row=fila_vacia, column=7, value=balance_inicial)
            ws.cell(row=fila_vacia, column=8, value=balance_final)
            ws.cell(row=fila_vacia, column=9, value=profit)

            # Guardar los cambios en el archivo Excel
            wb.save(nombre_archivo)
            break

        except Exception as e:
            # En caso de que ocurra una excepción
            print(Fore.RED + Back.WHITE +
                  f"Error durante la escritura del archivo de excel: {e}" + Back.BLACK + Fore.WHITE)
            logger.error(
                "¡Error %s al ejecutar la gaurdar los datos en el excel %s:\n%s\nCausa:%s\nTraceback:%s", type(e), var.direccion_excel, e, e.__cause__, e.__context__)
            attemps += 1

# Funcion para montar los datos necesarios para realizar la operacion para dejarla lista solo a la espera de la hora de ejecucion


async def montar_operacion(page, broker_url,  duracion_op: int, par_moneda: str, direccion: str, monto: int = 10, porcentaje: bool = True, gale: bool = False):
    """
    Monta una operación en el broker.

    Args:
        page (_type_): _description_
        broker_url (_type_): _description_
        duracion_op (int): _description_
        par_moneda (str): _description_
        direccion (str): _description_
        monto (int, optional): _description_. Defaults to 10.
        porcentaje (bool, optional): _description_. Defaults to True.
        gale (bool, optional): _description_. Defaults to False.
   """
    global operation_running
    operation_running = True

    global fechaHora_iniOperacion

    delta_seconds = (fechaHora_iniOperacion - datetime.now()).total_seconds()
    if delta_seconds < 0:
        return False

    attempts = 0
    while attempts < var.max_retry_attempts:
        try:
            if not await broker_check_and_login(page, var.quotex_email, var.quotex_password, var.broker_url_demo):
                print("Error al iniciar sesión en el broker")
                logger.error("Error al iniciar sesión en el broker")
                attempts += 1
                continue

            print(Back.BLACK + Fore.BLUE +
                  "Montando operación en el broker..." + Fore.WHITE + Style.NORMAL)
            logger.info("Montando operacion en el broker")

            etiqueta_tiempo = None

            etiqueta_tiempo = page.get_by_placeholder(":00:30")

            # PROCEDIMIENTO PARA ESTABLECER LA DURACION DE LA OPERACION
            etiqueta_tiempo_value = await etiqueta_tiempo.input_value()
            if etiqueta_tiempo_value != "00:01:00":  # si no esta marcada la opcion de tiempo en lugar de hora
                await page.reload(wait_until="load")  # recargamos la pagina
                # Espera a que la página se cargue completamente
                await page.wait_for_load_state()
                # esperamos a que cargue completamente
                # sleep(var.sleep_seconds_second)

            # hacemos click para seleccionar la duracion indicada en el mensaje de telegram
            await etiqueta_tiempo.click(delay=300)
            await asyncio.sleep(var.sleep_seconds_zero)
            tiempo = page.get_by_text(f"0{duracion_op}:00").first
            await tiempo.click(delay=300)

            # PROCEDMIENTO PARA ESTABLECER EL MONTO DE LA TRANSACCION
            # Seleccionamos el campo donde se agrega el valor
            valor_inversion = page.get_by_role("textbox").nth(1)
            await valor_inversion.fill(str(monto))  # marcamos 10

            tipo_inversion = page.locator(".section-deal__mobile-payout")

            # PROCEDMIENTO PARA CAMBIAR EL MODO DE INVERSION A PORCENTAJE
            if porcentaje:
                if not "percent" in await tipo_inversion.get_attribute("class"):
                    inversion_switch = page.get_by_text("Switch", exact=True)
                    await inversion_switch.click(delay=200)

            # PROCEDIMIENTO PARA INGRESAR EL PAR MONEDA DE LA OPERACION
            par_element = page.locator("#tab-active")
            await par_element.click(delay=300)
            # await asyncio.sleep(var.sleep_seconds_zero)
            buscador_par = page.get_by_placeholder("Buscar", exact=True)

            await buscador_par.fill(par_moneda)
            # sleep(var.sleep_seconds_zero)
            await page.locator(".assets-table__name ").first.click(delay=300)

            global init_balance
            init_balance = await page.locator(".usermenu__info-balance").inner_text()
            init_balance = float(init_balance[1:].replace(",", ""))

            global browser_page
            browser_page = page

            print(Fore.GREEN + Style.BRIGHT + "Operación montada en el broker " + Fore.WHITE + Style.NORMAL +
                  f'el {datetime.now().strftime("%d de %B de %Y a las %H:%M con %S segundos")}' + "\nA la espera de la hora de ejecución...")
            logger.info("Operacion montada en el broker")
            return True

        except Exception as to:

            print(f"¡Error al montar la operación en el broker!: {to}")
            logger.error(
                "¡Error %s al montar la operación en el broker!:\n%s\n Causa:%s\nTraceback:%s", type(to), to, to.__cause__, to.__context__)
            # recargamos la pagina en caso de que ya estemos en ella
            await page.reload(wait_until="load")
            # Espera a que la página se cargue completamente
            await page.wait_for_load_state()
            # await montar_operacion(page, broker_url, duracion_op, par_moneda, direccion, monto, porcentaje, gale)
            attempts += 1

    else:
        print(Back.BLACK + Fore.RED +
              "Maximo numero de reintentos alcanzado.\nNo se pudo montar la operación en el broker" + Fore.WHITE)
        operation_running = False
        return False


async def broker_check_and_login(page_context: Page, email: str, password: str, url: str = var.broker_url):
    """
    Asynchronously performs login in the broker using the provided email and password.

    Args:
    page_context (Page): The Playwright page object.
    email (str): The email for login.
    password (str): The password for login.
    url (str, optional): The URL for the broker page. Defaults to var.broker_url.

    Returns:
    bool: True if login is successful, False otherwise.
    """
    try:
        # Navegamos a la pagina del broker (Quotex)
        await page_context.goto(url)

        # Espera a que la página se cargue completamente
        await page_context.wait_for_load_state()
        # sleep(var.sleep_seconds_first)
        # Check if login form is visible
        if await page_context.get_by_role("textbox", name="Correo electrónico").is_visible():
            print(Back.BLACK + Fore.BLUE + "Realizando login en el broker...")

            # Fill in email and password fields
            email_field = page_context.get_by_role(
                "textbox", name="Correo electrónico")
            await email_field.fill(email)

            password_field = page_context.get_by_role(
                "textbox", name="Contraseña")
            await password_field.fill(password)

            # Click login button
            await page_context.get_by_role("button", name="Iniciar sesión").click(delay=300)
            # await asyncio.sleep(2)

            # Verify that the browser leaves th login page
            # await asyncio.sleep(var.sleep_seconds_first)
            for _ in range(6):
                if not await page_context.get_by_placeholder("Introduzca el código de 6 dígitos…").is_visible():
                    await asyncio.sleep(var.sleep_seconds_zero)
                else:
                    await asyncio.sleep(2)
                    break

            # Handle pin authentication if prompted
            if page_context.url == var.broker_url_login:
                pin_textbox = page_context.get_by_placeholder(
                    "Introduzca el código de 6 dígitos…")
                if await pin_textbox.is_visible():
                    print(
                        "El broker esta solicitando autenticacion mediante pin de acceso")
                    logger.info(
                        "Se solicito ingreso de pin de verificacion en el broker")

                    # TODO: agregar funcion para verificar si el codigo se introdujo correctamente y si no es asi, volver a pedirlo
                    error = True
                    while error:
                        code = input(Fore.YELLOW + Back.BLACK +
                                     "ingrese el codigo de acceso enviado a su correo registrado en el broker y pulse Enter: ")
                        await pin_textbox.fill(code)
                        await page_context.locator(".button--spaced").click(delay=300)
                        await page_context.wait_for_load_state()

                        await asyncio.sleep(6)
                        if await page_context.locator(".hint--danger").is_visible():
                            print(Fore.RED + Style.BRIGHT +
                                  "El pin ingresado es incorrecto" + Fore.WHITE)
                            error = True
                        else:
                            error = False
                            break

                    print("Codigo correcto")
                    logger.info("Se ingreso el codigo correctamente")
                else:
                    await broker_check_and_login(page_context, email, password, url)

            # Verify that the browser leaves the login page
            while True:
                if page_context.url != var.broker_url_login:
                    break
                await asyncio.sleep(var.sleep_seconds_zero)

                # await page_context.wait_for_url(url)
                # await page_context.wait_for_url(url)

            if page_context.url == var.broker_url:
                await page_context.goto(url)
                await page_context.wait_for_load_state()

            # Check if login is successful
            if page_context.url != url and page_context.url != var.broker_url_demo:
                logger.error(
                    "Hubo un error al intentar realizar el login en el broker")
                print(Fore.RED + Style.BRIGHT +
                      f"Hubo un error al intentar realizar el login en el broker el {datetime.now().strftime(var.formato_fecha)}" + Fore.WHITE)
                return False
            else:
                logger.info("Logeado exitosamente en el broker")
                print(
                    Fore.BLUE + f"Logeado exitosamente en el broker el {datetime.now().strftime(var.formato_fecha)}" + Fore.WHITE)
                return True
        else:
            print("Ya se encuentra logueado en el broker")
            logger.info("Ya se encuentra logueado en el broker")
            return True

    except Exception as e:
        logger.error(
            "Ocurrio un error al intentar logearse en el broker: " + str(e))
        return False

# Funcion para ejecutar la operacion. Pone en funcionamiento todo lo configurado en la funcion montar_operacion


async def ejecutar_operacion(page, direc_operacion, execution_datetime):
    """
    Executes a trading operation on the broker's page.
    Args:
        page (Page): The page object for the broker's website.
        operation_direction (str): The direction of the trading operation (call or put).
        execution_datetime (datetime): The date and time of the operation execution.
    Returns:
        dict: A dictionary containing the result of the operation.
    """
    intentos = 0
    resultado = 0
    hora_ini = execution_datetime.strftime("%H:%M:%S")
    broker_time = ""
    execution_completed = False
    execution_success = None
    result_tipe = None

    global init_balance
    global final_balance
    global operation_running

    # Esperamos a que se haga la hora de la ejecucion. Lo hacemos asi porque debemos
    # leer el reloj de la pagina del broker ya que este no esta sincronizado con el reloj del sistema
    while hora_ini != broker_time:
        broker_time = await page.locator(".server-time").inner_text()
        broker_time = broker_time[:8]
        if datetime.strptime(broker_time, "%H:%M:%S") > datetime.strptime(hora_ini, "%H:%M:%S"):
            print(Fore.YELLOW + "La hora de ejecución ya ha pasado" +
                  Fore.WHITE + "\nA la espera de una próxima senal")
            return True

    # while intentos < var.max_retry_attempts:
    try:
        execution_point = 0
        resultado = 0

        while not execution_completed:

            # PROCEDIMIENTO PARA SELECCIONAR EL TIPO DE OPERACION (CALL, PUT)
            if direc_operacion.lower().strip() == "put":
                await page.locator(".put-btn").click()
            elif direc_operacion.lower().strip() == 'call':
                await page.locator(".call-btn").click()

            # cronometro_start = time()

            fecha_inicio = datetime.now()
            print(Back.BLACK + Fore.BLUE + Style.BRIGHT +
                  f'Operación INICIADA el {fecha_inicio.strftime(var.formato_fecha)}\nEsperando resultados....')
            logger.info("operacion iniciada en el broker")

            execution_point = 1

            # PROCEDMIENTO PARA ESTABLECER EL MONTO DE LA TRANSACCION
            if intentos == 0:
                # Seleccionamos el campo donde se agrega el valor
                valor_inversion = page.get_by_role("textbox").nth(1)
                # marcamos 20 para que quede ya listo para el GALE
                await valor_inversion.fill(str(var.broker_amount_gale))
            elif intentos == 1:
                # Seleccionamos el campo donde se agrega el valor
                valor_inversion = page.get_by_role("textbox").nth(1)

                # marcamos 20 para que quede ya listo para el GALE
                await valor_inversion.fill(str(var.broker_amount_gale_2))

            execution_point = 2

            # Esperamos a que se ejecute la operacion, usualmente son 5 min
            # espera_fin = (var.sleep_1_minute * duracion)
            espera_fin = ""
            await asyncio.sleep(var.sleep_seconds_first)
            while espera_fin.strip() != "00:00":
                espera_fin = await page.locator(".trades-list-item__countdown").first.inner_text()

            # time_result = time()
            execution_point = 3
            res = await page.locator("#trade-item-open").inner_text()

            fecha_fin = datetime.now()
            print(Back.BLACK + Fore.BLUE + Style.BRIGHT +
                  f"Operación FINALIZADA el {fecha_fin.strftime('%d de %B de %Y a las %H:%M con %S segundos')}")
            logger.info("operacion en el broker finalizada")

            execution_point = 4

            resultado = float(res.split("\n")[3][:-2].replace(",", ""))
            # print(Fore.WHITE + Style.BRIGHT +
            #       f"Saldo inicial = {Fore.BLUE + Style.BRIGHT}${init_balance}")

            # sumamos el resultado de la operacion al saldo inicial

            if resultado > 0:
                print(Fore.WHITE + Style.BRIGHT +
                      f"Resultado de la operación: {Fore.GREEN + Style.BRIGHT}WIN.\tI")
                execution_completed = True
            else:
                print(Fore.WHITE + Style.BRIGHT +
                      f"Resultado de la operación: {Fore.RED + Style.BRIGHT}LOSE.\tO")
                intentos += 1
                if intentos <= 2:
                    execution_point += 1
                    print(Fore.BLUE + f"Se inició GALE {intentos}")
                else:
                    execution_completed = True

        operation_running = False

        execution_success = True
        logger.info("operacion exitosa en el broker")

        await asyncio.sleep(var.sleep_seconds_second)
        # await refresh_page(page)
        final_balance = await page.locator(".usermenu__info-balance").inner_text()
        final_balance = float(final_balance[1:].replace(",", "").strip())

        profit = 0

        if intentos >= 0:
            result_type = "D"
            # final_balance = (
            #     init_balance - round(init_balance*(var.broker_amount/100), 2))
        if intentos > 0 and intentos >= 1:
            result_type = "G1"
            # final_balance -= round(init_balance *
            #    (var.broker_amount_gale/100), 2)
        if intentos > 1 and intentos >= 2:
            result_type = "G2"
            # final_balance -= round(init_balance *
            #                        (var.broker_amount_gale_2/100), 2)
        if intentos > 2:
            result_type = "P"

        # final_balance += resultado

        profit = round(final_balance - init_balance, 2)
        final_balance = round(final_balance, 2)

        return_object = {
            "execution_success": execution_success,
            "init_balance": init_balance,
            "final_balance": final_balance,
            "profit": profit if profit > 0 else 0,
            "execution_date": execution_datetime.date(),
            "execution_time": execution_datetime.time(),
            'result_type': result_type,
            "direccion": direc_operacion.upper().strip(),
        }

        print(
            f"""{Back.BLACK + Fore.WHITE}
                ***RESUMEN***
            Balance Inicial = {Fore.BLUE + Style.BRIGHT}${return_object["init_balance"]}{Back.BLACK + Fore.WHITE}
            Balance Final = {Fore.BLUE + Style.BRIGHT}${return_object["final_balance"]}{Back.BLACK + Fore.WHITE}
            Tipo de resultado = {Fore.BLUE + Style.BRIGHT}{return_object["result_type"]}{Back.BLACK + Fore.WHITE}
            profit = {Fore.BLUE}${return_object["profit"]}{Back.BLACK + Fore.WHITE}
            """
        )
        return return_object

    except async_api.TimeoutError as e:
        # En caso de que ocurra una excepción
        print(Fore.RED + Back.WHITE +
              f"Error durante la ejecucion de la operación en el broker en la parte {execution_point}: {e}" + Back.BLACK + Fore.WHITE)
        logger.error(
            "¡Error %s al ejecutar la operación en el broker!:\n%s\n Causa:%s\nTraceback:%s", type(e), e, e.__cause__, e.__context__)

        if execution_point == 0:  # si es 0 quiere decir que el error ocurrio al pulsar los botones de call o put
            logger.error(
                "La operacion fallo al pulsar en los botones PUT o CALL")
            await ejecutar_operacion(page, direc_operacion, execution_datetime + timedelta(seconds=1))

        elif execution_point == 1:  # si parte es 1 quiere decir que el error ocurrio al seleccionar el campo para marcar el valor de la inversion
            logger.error(
                "La operacion ya esta corriendo en el broker pero fallo el intento de seleccionar el campo para ingresar el valor del monto de inversion")
            # await ejecutar_operacion(page, operacion)

        elif execution_point == 2:
            logger.error(
                "La operacion ya esta corriendo en el broker pero ocurrio un error al intentar cargar el valor de la inversion para la operacion del GALE."
            )

        elif execution_point == 3:
            logger.error(
                "La operacion termino pero fallo el intento de obtener el resultado")
            # await ejecutar_operacion(page, operacion)

        elif execution_point == 5:
            logger.error(
                "La operacion fallo cuando se estaba montando el GALE 1.")
            # await ejecutar_operacion(page, operacion)

        elif execution_point == 6:
            logger.error(
                "La operacion fallo cuando se estaba montando el GALE 2.")
            # await ejecutar_operacion(page, operacion)

        intentos += 1

        operation_running = False

    except Exception as e:
        operation_running = False

        print(Fore.RED + "Ocurrio un error inesperado: %s", e)
        logger.error("¡Error inesperado!: %s", e)
        intentos += 1


# Función asincrónica para lanzar el navegador


async def launch_browser():
    """Launches a new browser and sets up a new page for web scraping."""
    playwright = await async_playwright().start()
    # Launch the browser in headless mode
    browser = await playwright.firefox.launch(headless=False)
    context = await browser.new_context()
    global browser_page
    browser_page = await context.new_page()
    browser_page.set_default_timeout(timeout=10000)

    print("Navegador web abierto en modo oculto.")
    logger.info("Navegador web abierto")

    # Start the function to refresh the page in the background
    asyncio.create_task(refresh_page(browser_page))

    # try:
    #     # Wait until KeyboardInterrupt is applied
    #     await asyncio.Event().wait()
    # except KeyboardInterrupt:
    #     print("Closing the browser...")
    #     await browser.close()


# Función asincrónica para conectarse a Telegram y escuchar mensajes en un chat específico
async def launch_telegram(api_id, api_hash, chat_id):
    """_summary_

    Args:
        api_id (_type_): _description_
        api_hash (_type_): _description_
        chat_id (_type_): _description_
    """
    global telegram_client
     telegram_client = TelegramClient('data/session_name', api_id, api_hash)
    await telegram_client.start()

    # @telegram_client.on(events.NewMessage(chats=[chat_id,"me"], pattern=r'(?i).*Zona horaria: UTC')) #descomentar esta linea y comentar la siguiente para filtrar tambien los mensajes por chats
    @telegram_client.on(events.NewMessage(pattern=r'(?i).*Zona horaria: UTC'))
    async def handle_new_message(event):
        global browser_page
        global browser_page_gale

        # global operation_running
        # operation_running = True

        global broker_signal_message
        broker_signal_message = event.message.text

        print(
            Fore.CYAN + f"\n\nNuevo mensaje en el chat {chat_id}. Recibido el {datetime.now().strftime(var.formato_fecha)}")
        logger.info("LLego una señal por telegram")

        text = list(event.message.text.split("\n"))

        global duracion
        duracion = int(text[1].split(' ')[0][-1])
        # duracion = 1 #duracion para ahcer prueas, se debe comentar esta linea en la ejecucion real y descomentar la linea anterior

        global par
        par = text[2].split(";")[0]

        global fechaHora_iniOperacion
        fechaHora_iniOperacion = ""

        global fechaHora_finOperacion
        fechaHora_finOperacion = ""

        global operacion
        operacion = ""

        try:
            fechaHora_iniOperacion = (datetime.strptime(
                text[2].split(";")[1], "%H:%M") - timedelta(hours=1))
            operacion = "CALL" if "CALL" in text[2].split(";")[2] else "PUT"
        except Exception:
            fechaHora_iniOperacion = (datetime.strptime(text[2].split(
                ";")[1].split(" ")[0], "%H:%M") - timedelta(hours=1))
            operacion = True if "CALL" in text[2].split(";")[1].split(" ")[
                1] else False

        fecha_actual = datetime.today()
        fechaHora_iniOperacion = fechaHora_iniOperacion.replace(
            year=fecha_actual.year, month=fecha_actual.month, day=fecha_actual.day, second=0, microsecond=0)
        fechaHora_finOperacion = fechaHora_iniOperacion + \
            timedelta(minutes=duracion)

        print(Fore.BLUE + Style.BRIGHT + Back.BLACK +
              "*************Datos para la operación*************")
        print(
            Fore.WHITE + f"Duracion de la operación: {Fore.BLUE + Style.BRIGHT}{duracion} minutos")
        print(Fore.WHITE +
              f"Par de la operación: {Fore.BLUE + Style.BRIGHT}{par}")
        print(
            Fore.WHITE + f'Fecha y hora de inicio de la operación: {Fore.BLUE + Style.BRIGHT}{fechaHora_iniOperacion.strftime("%d de %B de %Y a las %H:%M")}')
        print(
            Fore.WHITE + f'Fecha y hora de fin de la operación: {Fore.BLUE + Style.BRIGHT}{fechaHora_finOperacion.strftime("%d de %B de %Y a las %H:%M")}')
        print(
            Fore.WHITE + f"Direccion de la operación: {Fore.BLUE + Style.BRIGHT}{operacion}")

        logger.info(
            "Datos de la operación:\n%s minutos\n%s\n%s\n%s\n%s", duracion, par, fechaHora_iniOperacion.strftime(var.formato_fecha), fechaHora_finOperacion.strftime(var.formato_fecha), operacion)

        # hora de ejecucion del proceso de montaje de la operacion (3 segundos a partir de ahora)
        un_poco_despues = datetime.now() + timedelta(seconds=3)

        op_montada = await montar_operacion(browser_page, var.broker_url_demo, duracion, par, operacion, var.broker_amount, var.broker_percent)

        if op_montada:
            result = await ejecutar_operacion(browser_page, operacion, fechaHora_iniOperacion)
            agregar_a_excel(
                result["execution_date"], result["execution_time"], duracion, result["result_type"],
                par, result["direccion"], result["init_balance"], result["final_balance"], result["profit"]
            )
            print(Fore.WHITE + "Operacion ejecutada. A la espera de una próxima señal")
        else:
            print(Fore.YELLOW + "A la espera de una próxima señal" + Fore.WHITE)

    print("Conectado a Telegram. Escuchando mensajes entrantes que cumplan con el patrón")
    logger.info("Conexion a telegram exitosa")

    await telegram_client.run_until_disconnected()


# Función asincrónica para programar la ejecución de otra función en una fecha y hora específicas
async def schedule_function(execution_datetime: datetime, function_to_execute, *args, **kwargs):
    """_summary_

    Args:
        execution_datetime (_type_): _description_
        function_to_execute (_type_): _description_
    """
    # este margen de error lo ponemos por si la operacion tarda un poco mas de lo previsto en realizarse
    delta_seconds = (execution_datetime - datetime.now()).total_seconds()
    if delta_seconds > 0:
        # if execution_datetime + 15 > datetime.now():
        await asyncio.sleep(delta_seconds)
        await function_to_execute(*args, **kwargs)
        return True
    else:
        global operation_running
        operation_running = False
        print(Fore.YELLOW + "La fecha indicada ya ha pasado.\n" +
              Fore.WHITE + "A la espera de un nuevo mensaje.")
        logger.error(
            "La fecha de la operacion ya ha pasado\nA la espera de un nuevo mensaje.")
        return False

# Función principal que implementa las tres funciones anteriores


async def main():
    """
    This function is the main entry point of the program.

    It initializes the Telegram configuration parameters and launches the browser and Telegram concurrently.
    """
    # Telegram configuration parameters
    api_id = var.telegram_api_id
    api_hash = var.telegram_api_hash
    chat_id = var.telegram_group_username

    # Initialize the colorama module
    init()

    # Launch the browser and Telegram concurrently
    await asyncio.gather(
        launch_browser(),  # Launch the browser
        launch_telegram(api_id, api_hash, chat_id)  # Launch Telegram
    )


if __name__ == "__main__":
    asyncio.run(main())
